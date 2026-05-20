"""
Price Monitor — track product prices across e-commerce stores.

A small CLI tool that:
  1. Reads a list of product URLs + CSS selectors from products.yaml
  2. Fetches each page, extracts the current price
  3. Appends the result to history.csv (so you build a price timeline)
  4. Compares to the previous snapshot and flags drops / rises
  5. Writes an Excel workbook (report.xlsx) and an HTML report (report.html)

Demo target: books.toscrape.com — a site explicitly built for scraping
practice. Swap the URLs/selectors in products.yaml to track any store
that allows scraping (always check the site's robots.txt and ToS first).

Usage:
    python price_monitor.py                 # run once
    python price_monitor.py --notify        # also print a Slack/Discord-ready summary
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import requests
import yaml
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
PRODUCTS_FILE = ROOT / "products.yaml"
HISTORY_FILE = ROOT / "history.csv"
EXCEL_FILE = ROOT / "report.xlsx"
HTML_FILE = ROOT / "report.html"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; PriceMonitor/1.0; "
        "+https://example.com/price-monitor)"
    )
}
REQUEST_TIMEOUT = 15
SLEEP_BETWEEN = 1.0  # be polite — one request per second


@dataclass
class Product:
    name: str
    url: str
    selector: str
    target_price: float | None = None  # alert if current <= target


@dataclass
class Snapshot:
    name: str
    url: str
    price: float | None
    currency: str
    checked_at: str
    error: str | None = None


# ---------------------------------------------------------------------------
# Loading & fetching
# ---------------------------------------------------------------------------

def load_products() -> list[Product]:
    if not PRODUCTS_FILE.exists():
        sys.exit(f"products.yaml not found at {PRODUCTS_FILE}")
    raw = yaml.safe_load(PRODUCTS_FILE.read_text(encoding="utf-8"))
    return [Product(**item) for item in raw["products"]]


PRICE_RE = re.compile(r"(?P<cur>[£$€])?\s*(?P<num>\d[\d,]*\.?\d*)")


def parse_price(text: str) -> tuple[float | None, str]:
    """Pull a number and (optional) currency symbol out of a price string."""
    if not text:
        return None, ""
    match = PRICE_RE.search(text.replace("\xa0", " "))
    if not match:
        return None, ""
    number = match.group("num").replace(",", "")
    try:
        return float(number), match.group("cur") or ""
    except ValueError:
        return None, match.group("cur") or ""


def fetch_price(product: Product) -> Snapshot:
    now = datetime.now().isoformat(timespec="seconds")
    try:
        response = requests.get(product.url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
    except requests.RequestException as exc:
        return Snapshot(product.name, product.url, None, "", now, error=str(exc))

    soup = BeautifulSoup(response.text, "html.parser")
    node = soup.select_one(product.selector)
    if node is None:
        return Snapshot(
            product.name, product.url, None, "", now,
            error=f"selector '{product.selector}' returned nothing",
        )

    price, currency = parse_price(node.get_text(strip=True))
    if price is None:
        return Snapshot(
            product.name, product.url, None, currency, now,
            error=f"could not parse a number from '{node.get_text(strip=True)}'",
        )
    return Snapshot(product.name, product.url, price, currency, now)


# ---------------------------------------------------------------------------
# History
# ---------------------------------------------------------------------------

HISTORY_HEADER = ["checked_at", "name", "url", "price", "currency", "error"]


def append_history(snapshots: Iterable[Snapshot]) -> None:
    new_file = not HISTORY_FILE.exists()
    with HISTORY_FILE.open("a", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        if new_file:
            writer.writerow(HISTORY_HEADER)
        for snap in snapshots:
            writer.writerow([
                snap.checked_at, snap.name, snap.url,
                f"{snap.price:.2f}" if snap.price is not None else "",
                snap.currency, snap.error or "",
            ])


def previous_price(name: str, before_iso: str) -> float | None:
    """Return the price from the prior *run* — the most recent snapshot
    strictly older than this run AND from a different calendar date,
    so re-running the script on the same day still compares to yesterday."""
    if not HISTORY_FILE.exists():
        return None
    run_date = before_iso[:10]
    last: float | None = None
    with HISTORY_FILE.open("r", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            if row["name"] != name:
                continue
            if row["checked_at"] >= before_iso:
                continue
            if row["checked_at"][:10] == run_date:
                continue  # skip earlier runs from the same day
            if row["price"]:
                last = float(row["price"])
    return last


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------

def write_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Prices"

    headers = ["Product", "Current", "Previous", "Change", "Change %", "Target", "Status", "URL"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    drop_fill = PatternFill("solid", fgColor="DCFCE7")  # green
    rise_fill = PatternFill("solid", fgColor="FEE2E2")  # red
    hit_fill = PatternFill("solid", fgColor="FEF9C3")   # yellow — target hit

    for row in rows:
        ws.append([
            row["name"],
            row["current"],
            row["previous"],
            row["change"],
            f"{row['change_pct']:.1f}%" if row["change_pct"] is not None else "",
            row["target"],
            row["status"],
            row["url"],
        ])
        excel_row = ws[ws.max_row]
        if row["status"] == "DROP":
            for cell in excel_row:
                cell.fill = drop_fill
        elif row["status"] == "RISE":
            for cell in excel_row:
                cell.fill = rise_fill
        if row["target_hit"]:
            for cell in excel_row:
                cell.fill = hit_fill

    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 22
    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)


def write_html(rows: list[dict], generated_at: str) -> None:
    css = """
    body { font-family: -apple-system, Segoe UI, Roboto, sans-serif; background: #f5f6fa;
           color: #1f2937; padding: 32px; max-width: 1100px; margin: 0 auto; }
    h1 { margin: 0 0 4px; font-size: 28px; }
    .meta { color: #6b7280; margin-bottom: 24px; font-size: 14px; }
    table { width: 100%; border-collapse: collapse; background: #fff;
            border-radius: 12px; overflow: hidden; box-shadow: 0 4px 18px rgba(0,0,0,0.06); }
    th, td { text-align: left; padding: 12px 14px; font-size: 14px;
             border-bottom: 1px solid #eef0f4; }
    th { background: #1f2937; color: #fff; font-weight: 600; }
    tr:last-child td { border-bottom: none; }
    .drop  { background: #ecfdf5; }
    .rise  { background: #fef2f2; }
    .hit   { background: #fefce8; }
    .badge { display: inline-block; padding: 3px 9px; border-radius: 999px;
             font-size: 12px; font-weight: 600; }
    .badge--drop { background: #10b981; color: #fff; }
    .badge--rise { background: #ef4444; color: #fff; }
    .badge--same { background: #e5e7eb; color: #374151; }
    .badge--new  { background: #6366f1; color: #fff; }
    .badge--hit  { background: #eab308; color: #fff; margin-left: 6px; }
    a { color: #2563eb; text-decoration: none; }
    a:hover { text-decoration: underline; }
    .summary { display: flex; gap: 16px; flex-wrap: wrap; margin: 0 0 24px; }
    .card { background: #fff; padding: 16px 20px; border-radius: 12px;
            box-shadow: 0 4px 18px rgba(0,0,0,0.05); flex: 1 1 180px; }
    .card b { display: block; font-size: 22px; }
    .card span { color: #6b7280; font-size: 13px; }
    """

    total = len(rows)
    drops = sum(1 for r in rows if r["status"] == "DROP")
    rises = sum(1 for r in rows if r["status"] == "RISE")
    hits = sum(1 for r in rows if r["target_hit"])

    body_rows = []
    for row in rows:
        row_cls = ""
        if row["status"] == "DROP":
            row_cls = "drop"
        elif row["status"] == "RISE":
            row_cls = "rise"
        if row["target_hit"]:
            row_cls = "hit"

        status_badge = {
            "DROP": '<span class="badge badge--drop">↓ drop</span>',
            "RISE": '<span class="badge badge--rise">↑ rise</span>',
            "SAME": '<span class="badge badge--same">same</span>',
            "NEW":  '<span class="badge badge--new">new</span>',
            "N/A":  '<span class="badge badge--same">n/a</span>',
        }.get(row["status"], "")
        if row["target_hit"]:
            status_badge += '<span class="badge badge--hit">target hit</span>'

        previous = f"{row['previous']:.2f}" if row["previous"] is not None else "—"
        change = f"{row['change']:+.2f}" if row["change"] is not None else "—"
        change_pct = f"{row['change_pct']:+.1f}%" if row["change_pct"] is not None else "—"
        target = f"{row['target']:.2f}" if row["target"] is not None else "—"
        current = f"{row['current']:.2f}" if row["current"] is not None else "(error)"

        body_rows.append(
            f"<tr class='{row_cls}'>"
            f"<td><b>{row['name']}</b></td>"
            f"<td>{current}</td>"
            f"<td>{previous}</td>"
            f"<td>{change}</td>"
            f"<td>{change_pct}</td>"
            f"<td>{target}</td>"
            f"<td>{status_badge}</td>"
            f"<td><a href='{row['url']}' target='_blank'>open ↗</a></td>"
            f"</tr>"
        )

    html = f"""<!doctype html>
<html lang="en"><head>
<meta charset="utf-8"><title>Price Monitor Report</title>
<style>{css}</style>
</head><body>
  <h1>Price Monitor — daily report</h1>
  <p class="meta">Generated {generated_at}</p>
  <div class="summary">
    <div class="card"><b>{total}</b><span>products tracked</span></div>
    <div class="card"><b>{drops}</b><span>price drops</span></div>
    <div class="card"><b>{rises}</b><span>price rises</span></div>
    <div class="card"><b>{hits}</b><span>target prices hit</span></div>
  </div>
  <table>
    <thead><tr>
      <th>Product</th><th>Current</th><th>Previous</th><th>Change</th>
      <th>Change %</th><th>Target</th><th>Status</th><th>Link</th>
    </tr></thead>
    <tbody>{''.join(body_rows)}</tbody>
  </table>
</body></html>"""
    HTML_FILE.write_text(html, encoding="utf-8")


# ---------------------------------------------------------------------------
# Main flow
# ---------------------------------------------------------------------------

def build_rows(products: list[Product], snapshots: list[Snapshot], run_started: str) -> list[dict]:
    rows: list[dict] = []
    for product, snap in zip(products, snapshots):
        previous = previous_price(product.name, before_iso=run_started)
        current = snap.price
        change = change_pct = None
        if current is not None and previous is not None:
            change = current - previous
            change_pct = (change / previous) * 100 if previous else None

        if snap.error or current is None:
            status = "N/A"
        elif previous is None:
            status = "NEW"
        elif change is not None and change < -0.005:
            status = "DROP"
        elif change is not None and change > 0.005:
            status = "RISE"
        else:
            status = "SAME"

        target_hit = (
            product.target_price is not None
            and current is not None
            and current <= product.target_price
        )

        rows.append({
            "name": product.name,
            "url": product.url,
            "current": current,
            "previous": previous,
            "change": change,
            "change_pct": change_pct,
            "target": product.target_price,
            "status": status,
            "target_hit": target_hit,
            "error": snap.error,
        })
    return rows


def notify_summary(rows: list[dict]) -> str:
    """Build a short summary message for Slack/Discord/email."""
    interesting = [r for r in rows if r["status"] in ("DROP",) or r["target_hit"]]
    if not interesting:
        return "Price Monitor: no notable changes today."
    lines = ["*Price Monitor - today's highlights*"]
    for row in interesting:
        tag = "[TARGET HIT]" if row["target_hit"] else "[DROP]"
        change = f" ({row['change_pct']:+.1f}%)" if row["change_pct"] is not None else ""
        lines.append(f"- {tag} {row['name']}: {row['current']:.2f}{change}")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Track product prices across stores.")
    parser.add_argument("--notify", action="store_true",
                        help="Print a notification summary to stdout.")
    args = parser.parse_args()

    products = load_products()
    run_started = datetime.now().isoformat(timespec="seconds")

    print(f"Checking {len(products)} products...")
    snapshots: list[Snapshot] = []
    for product in products:
        snap = fetch_price(product)
        flag = "OK" if snap.price is not None else "ERR"
        price = f"{snap.price:.2f}" if snap.price is not None else snap.error
        print(f"  [{flag}] {product.name}: {price}")
        snapshots.append(snap)
        time.sleep(SLEEP_BETWEEN)

    append_history(snapshots)
    rows = build_rows(products, snapshots, run_started)
    write_excel(rows)
    write_html(rows, run_started)

    print()
    print(f"Excel report: {EXCEL_FILE.name}")
    print(f"HTML report:  {HTML_FILE.name}")

    if args.notify:
        print()
        print(notify_summary(rows))


if __name__ == "__main__":
    main()
